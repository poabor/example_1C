import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def fill_empty_cells(input_file, input_sheet, columns_to_fill):
    """
    Заполняет пустые ячейки значениями сверху
    """
    # Чтение данных
    df = pd.read_excel(input_file, sheet_name=input_sheet)
    
    # Заполнение указанных столбцов
    for col in columns_to_fill:
        if col in df.columns:
            df[col] = df[col].ffill()
    
    name_output_file = 'output_' + input_file # название выходного файла
    name_output_sheet = 'output_' + input_sheet # название выходного листа

    # Сохранение с автошириной
    with pd.ExcelWriter(name_output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=name_output_sheet, index=False)
        
        worksheet = writer.sheets[name_output_sheet]
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            if max_length > 0:
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    print(f"Готово! Файл сохранен как '{name_output_file}'")

# Использование
fill_empty_cells(
    input_file='tblPurifications.xlsx', # имя обрабатываемого файла
    input_sheet='ТехнСерии_2',  # лист для чтения
    columns_to_fill=['Каталожный №'] # название колонки, которую надо заполнить. берется из первой строки колонки
)